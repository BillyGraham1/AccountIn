"""
Accounting System — General Journal → Ledger → Trial Balance → Financial Statements
=====================================================================================
A Streamlit double-entry bookkeeping app for a sole proprietorship.

Pages
-----
Input Journal | View Journal | View Ledger | Trial Balance |
Financial Statements (Income Statement, Statement of Owner's Equity, Balance Sheet) |
Export Full Cycle | Chart of Accounts

The app opens straight into Input Journal — there is no separate Dashboard.

Design notes
------------
* Every account in the Chart of Accounts carries an explicit `type`
  (Asset / Liability / Equity / Revenue / Expense), a `subtype` used for
  statement grouping, and a `contra` flag. Nothing about account
  classification is guessed from the account code any more.
* All balances are computed from raw Debit/Credit postings with a single
  consistent convention: account_net() = sum(Debit) - sum(Credit). Every
  report derives its numbers from that one function, so Trial Balance,
  Income Statement, Statement of Owner's Equity and Balance Sheet can never
  drift out of sync with each other.
* The Statement of Owner's Equity / Balance Sheet always tie out exactly,
  for ANY period start date, because "Capital, Beginning" is computed as
  cumulative contributions + cumulative net income - cumulative drawings
  from inception (not just the literal Capital account balance) — see
  compute_owners_equity() for the derivation.
* Currency (Rp / $) is a *display* setting only — it changes labels,
  number formatting and Excel number formats, not the stored values.
"""

import io
import json
import uuid
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st

# Some pandas 3.x + pyarrow combinations have a real crash (SIGSEGV, not just
# an exception) in the Arrow-backed string array code path — reachable via
# perfectly ordinary calls like .unique()/.nunique()/sort_values() — when run
# repeatedly inside Streamlit's threaded script runner. Forcing plain object
# dtype for strings sidesteps that whole class of crash; it has no visible
# effect on correctness, just how strings are stored internally.
pd.set_option("future.infer_string", False)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    _OPENPYXL_AVAILABLE = False

# ============================================================================
# PAGE CONFIG
# ============================================================================

st.set_page_config(
    page_title="Accounting System",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================================
# THEME / CSS
# Clean, modern identity carried over from the original app: a light
# background, a single blue accent, plain sans-serif type, and light-bordered
# cards — no textured paper, no serif display face. Numbers keep a
# monospaced face purely so columns of figures line up neatly.
# All selectors target stable data-testid hooks, not hashed Streamlit
# classnames (those change between versions and silently stop matching).
# ============================================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

:root {
    --paper: #FFFFFF;
    --ink: #1C2B33;
    --ink-soft: #5B6B73;
    --accent: #1F77B4;
    --accent-dark: #145A8A;
    --accent-tint: #EAF3FB;
    --gold: #B8873B;
    --gold-tint: #F7EEDD;
    --burgundy: #B00020;
    --line: #E2E8F0;
    --card: #FFFFFF;
}

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
* { color: var(--ink); }

.stApp { background-color: var(--paper); }

h1, h2, h3 { font-family: 'Inter', sans-serif; color: var(--ink); font-weight: 700 !important; }
h1 { color: var(--accent); padding-bottom: 0.5rem; margin-bottom: 0.75rem !important; }
h2 { color: var(--ink); padding-top: 0.25rem; }
h3 { color: var(--accent-dark); }

p, span, label, div { font-family: 'Inter', sans-serif; }

/* ---------- Sidebar ---------- */
[data-testid="stSidebar"] {
    background: #F7F9FB;
    border-right: 1px solid var(--line);
}
[data-testid="stSidebar"] hr { border-top: 1px solid var(--line); margin: 0.9rem 0; }
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 {
    color: var(--accent);
}

[data-testid="stSidebar"] button {
    border-radius: 8px !important;
    border: 1px solid var(--line) !important;
    background: #FFFFFF !important;
    font-weight: 500 !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] button p { text-align: left !important; }
[data-testid="stSidebar"] button:hover {
    background: var(--accent-tint) !important;
    border-color: var(--accent) !important;
}
[data-testid="stSidebar"] button[kind="primary"] {
    background: var(--accent) !important;
    border-color: var(--accent) !important;
}
[data-testid="stSidebar"] button[kind="primary"] * { color: #FFFFFF !important; }

.nav-eyebrow {
    font-family: 'Inter', sans-serif;
    text-transform: uppercase;
    letter-spacing: 0.09em;
    font-size: 0.68rem;
    color: var(--ink-soft) !important;
    margin: 0.9rem 0 0.35rem 0.1rem;
    font-weight: 600;
}

/* ---------- Metrics ---------- */
[data-testid="stMetric"] {
    background: var(--card);
    border: 1px solid var(--line);
    border-left: 4px solid var(--accent);
    border-radius: 10px;
    padding: 0.85rem 1rem;
    box-shadow: 0 1px 2px rgba(28,43,51,0.05);
}
[data-testid="stMetricValue"] {
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 1.25rem !important;
}
[data-testid="stMetricLabel"] { color: var(--ink-soft) !important; font-weight: 600 !important; font-size: 0.8rem !important; }

/* ---------- Buttons (main area) ---------- */
.stButton>button {
    border-radius: 8px;
    font-weight: 600;
}
div[data-testid="stAppViewContainer"] button[kind="primary"] {
    background: var(--accent) !important;
    border-color: var(--accent) !important;
}
div[data-testid="stAppViewContainer"] button[kind="primary"]:hover {
    background: var(--accent-dark) !important;
    border-color: var(--accent-dark) !important;
}

/* ---------- Tables ---------- */
[data-testid="stDataFrame"], [data-testid="stTable"] {
    border-radius: 8px;
    overflow: hidden;
    border: 1px solid var(--line);
}
[data-testid="stDataFrame"] * { font-family: 'IBM Plex Mono', monospace !important; font-size: 13.5px; }

/* ---------- Misc ---------- */
hr { border-top: 1px solid var(--line); }
div[data-testid="stAlert"] { border-radius: 8px; }
[data-testid="stTabs"] button[role="tab"] { font-weight: 600; }

.ledger-card {
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 1.1rem 1.4rem;
    box-shadow: 0 1px 3px rgba(28,43,51,0.05);
    margin-bottom: 0.9rem;
}
.ledger-eyebrow {
    font-family: 'Inter', sans-serif;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    font-size: 0.72rem;
    color: var(--accent);
    font-weight: 700;
}
.badge-debit {
    display:inline-block; padding: 2px 10px; border-radius: 999px;
    background: var(--accent-tint); color: var(--accent-dark); font-weight:700; font-size:0.76rem;
}
.badge-credit {
    display:inline-block; padding: 2px 10px; border-radius: 999px;
    background: var(--gold-tint); color: #7A5A1E; font-weight:700; font-size:0.76rem;
}
.status-balanced { color: var(--accent-dark); font-weight:700; }
.status-unbalanced { color: var(--burgundy); font-weight:700; }
.statement-line { display:flex; justify-content:space-between; padding: 0.3rem 0; border-bottom: 1px dashed var(--line); font-family:'IBM Plex Mono',monospace; }
.statement-total { display:flex; justify-content:space-between; padding: 0.5rem 0; border-top: 2px solid var(--ink); border-bottom: 4px double var(--ink); font-family:'IBM Plex Mono',monospace; font-weight:700; font-size:1.05rem; margin-top:0.25rem;}
.statement-subtotal { display:flex; justify-content:space-between; padding: 0.35rem 0; border-top: 1px solid var(--line); font-family:'IBM Plex Mono',monospace; font-weight:700; }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# CONSTANTS — Chart of Accounts schema
# ============================================================================

APP_DIR = Path(__file__).resolve().parent

ACCOUNT_TYPES = ["Asset", "Liability", "Equity", "Revenue", "Expense"]

SUBTYPES_BY_TYPE = {
    "Asset": ["Current Asset", "Fixed Asset", "Other Asset"],
    "Liability": ["Current Liability", "Long-Term Liability"],
    "Equity": ["Capital", "Drawing"],
    "Revenue": ["Operating Revenue", "Other Revenue"],
    "Expense": ["Operating Expense", "Other Expense"],
}

TYPE_ICON = {"Asset": "🏦", "Liability": "💳", "Equity": "👤", "Revenue": "💰", "Expense": "🧾"}

BASE_NORMAL_BALANCE = {
    "Asset": "Debit", "Liability": "Credit", "Equity": "Credit",
    "Revenue": "Credit", "Expense": "Debit",
}

DEFAULT_COA = {
    "101": {"name": "Cash", "type": "Asset", "subtype": "Current Asset", "contra": False},
    "102": {"name": "Accounts Receivable", "type": "Asset", "subtype": "Current Asset", "contra": False},
    "103": {"name": "Inventory", "type": "Asset", "subtype": "Current Asset", "contra": False},
    "104": {"name": "Prepaid Expenses", "type": "Asset", "subtype": "Current Asset", "contra": False},
    "105": {"name": "Office Supplies", "type": "Asset", "subtype": "Current Asset", "contra": False},
    "150": {"name": "Equipment", "type": "Asset", "subtype": "Fixed Asset", "contra": False},
    "151": {"name": "Accumulated Depreciation - Equipment", "type": "Asset", "subtype": "Fixed Asset", "contra": True},
    "201": {"name": "Accounts Payable", "type": "Liability", "subtype": "Current Liability", "contra": False},
    "202": {"name": "Notes Payable", "type": "Liability", "subtype": "Current Liability", "contra": False},
    "203": {"name": "Salaries Payable", "type": "Liability", "subtype": "Current Liability", "contra": False},
    "204": {"name": "Interest Payable", "type": "Liability", "subtype": "Current Liability", "contra": False},
    "301": {"name": "Owner's Capital", "type": "Equity", "subtype": "Capital", "contra": False},
    "302": {"name": "Owner's Drawings", "type": "Equity", "subtype": "Drawing", "contra": False},
    "401": {"name": "Service Revenue", "type": "Revenue", "subtype": "Operating Revenue", "contra": False},
    "402": {"name": "Sales Revenue", "type": "Revenue", "subtype": "Operating Revenue", "contra": False},
    "501": {"name": "Rent Expense", "type": "Expense", "subtype": "Operating Expense", "contra": False},
    "502": {"name": "Salaries Expense", "type": "Expense", "subtype": "Operating Expense", "contra": False},
    "503": {"name": "Utilities Expense", "type": "Expense", "subtype": "Operating Expense", "contra": False},
    "504": {"name": "Depreciation Expense", "type": "Expense", "subtype": "Operating Expense", "contra": False},
    "505": {"name": "Supplies Expense", "type": "Expense", "subtype": "Operating Expense", "contra": False},
    "506": {"name": "Interest Expense", "type": "Expense", "subtype": "Other Expense", "contra": False},
}

JOURNAL_COLUMNS = [
    "Date", "Transaction_ID", "Line_ID", "Account_Title",
    "Post_Ref", "Explanation", "Debit", "Credit",
]

CURRENCY_META = {
    "Rp": {"symbol": "Rp", "decimals": 0},
    "$": {"symbol": "$", "decimals": 2},
}

# ============================================================================
# DATA PERSISTENCE LAYER
# ============================================================================

class DataManager:
    """Handles all data storage and retrieval. Files live next to this
    script (not the current working directory), so the app finds the same
    data no matter where `streamlit run` was launched from."""

    def __init__(self):
        self.data_dir = APP_DIR / "accounting_data"
        self.data_dir.mkdir(exist_ok=True)
        self.chart_file = self.data_dir / "chart_of_accounts.json"
        self.journal_file = self.data_dir / "general_journal.csv"
        self.settings_file = self.data_dir / "settings.json"

    @staticmethod
    def _backup(path: Path):
        """Rename an incompatible/legacy file out of the way instead of
        silently deleting it, then start fresh."""
        if path.exists():
            try:
                backup = path.with_suffix(path.suffix + ".bak")
                path.replace(backup)
            except Exception:
                pass

    # -- Chart of Accounts --------------------------------------------------
    def load_chart_of_accounts(self):
        if self.chart_file.exists():
            try:
                with open(self.chart_file, "r") as f:
                    raw = json.load(f)
                valid = (
                    isinstance(raw, dict) and len(raw) > 0
                    and all(
                        isinstance(v, dict) and {"name", "type", "subtype"} <= v.keys()
                        and v["type"] in ACCOUNT_TYPES
                        for v in raw.values()
                    )
                )
                if valid:
                    for v in raw.values():
                        v.setdefault("contra", False)
                    return raw
            except Exception:
                pass
            self._backup(self.chart_file)
        chart = {code: dict(info) for code, info in DEFAULT_COA.items()}
        self.save_chart_of_accounts(chart)
        return chart

    def save_chart_of_accounts(self, chart):
        with open(self.chart_file, "w") as f:
            json.dump(chart, f, indent=2)

    # -- General Journal ------------------------------------------------------
    def load_journal(self):
        if self.journal_file.exists():
            try:
                df = pd.read_csv(self.journal_file, dtype={"Post_Ref": str})
                if set(JOURNAL_COLUMNS) <= set(df.columns):
                    df["Date"] = pd.to_datetime(df["Date"])
                    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0.0)
                    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0.0)
                    # Force plain python 'object' dtype for text columns. Newer
                    # pandas defaults to an Arrow-backed string dtype when
                    # reading CSVs, and that dtype has a real crash (SIGSEGV)
                    # in .unique()/.nunique() in this environment — object
                    # dtype sidesteps it entirely and is what a fresh empty
                    # DataFrame uses anyway, so behaviour stays consistent.
                    for col in ("Transaction_ID", "Line_ID", "Account_Title", "Post_Ref", "Explanation"):
                        df[col] = df[col].astype(object).where(df[col].notna(), "")
                    return df[JOURNAL_COLUMNS]
            except Exception:
                pass
            self._backup(self.journal_file)
        return pd.DataFrame(columns=JOURNAL_COLUMNS)

    def save_journal(self, df: pd.DataFrame):
        df.to_csv(self.journal_file, index=False)

    # -- Settings (currency preference, etc.) --------------------------------
    def load_settings(self):
        if self.settings_file.exists():
            try:
                with open(self.settings_file, "r") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"currency": "Rp"}

    def save_settings(self, settings: dict):
        with open(self.settings_file, "w") as f:
            json.dump(settings, f, indent=2)


data_mgr = DataManager()

# ============================================================================
# SESSION STATE INITIALIZATION
# ============================================================================

if "chart_of_accounts" not in st.session_state:
    st.session_state.chart_of_accounts = data_mgr.load_chart_of_accounts()

if "journal_df" not in st.session_state:
    st.session_state.journal_df = data_mgr.load_journal()

if "app_settings" not in st.session_state:
    st.session_state.app_settings = data_mgr.load_settings()

if "currency" not in st.session_state:
    st.session_state.currency = st.session_state.app_settings.get("currency", "Rp")

if "nav_page" not in st.session_state:
    st.session_state.nav_page = "Input Journal"

if "entry_form_key" not in st.session_state:
    st.session_state.entry_form_key = 0

if "confirm_delete_all" not in st.session_state:
    st.session_state.confirm_delete_all = False

if "confirm_delete_txn" not in st.session_state:
    st.session_state.confirm_delete_txn = None

if "confirm_remove_acct" not in st.session_state:
    st.session_state.confirm_remove_acct = None

# ============================================================================
# UTILITY FUNCTIONS — Chart of Accounts helpers
# ============================================================================

def coa() -> dict:
    return st.session_state.chart_of_accounts

def account_info(code) -> dict:
    return coa().get(str(code), {
        "name": "Unknown Account", "type": "Asset", "subtype": "Current Asset", "contra": False,
    })

def account_name(code) -> str:
    return account_info(code)["name"]

def account_normal_balance(code) -> str:
    """Which side (Debit/Credit) is this account's natural increasing side."""
    info = account_info(code)
    base = BASE_NORMAL_BALANCE.get(info["type"], "Debit")
    if info.get("subtype") == "Drawing":
        base = "Debit"
    if info.get("contra"):
        base = "Credit" if base == "Debit" else "Debit"
    return base

def sorted_account_codes() -> list:
    return sorted(coa().keys(), key=lambda c: (len(c), c))

def account_select_options(codes=None):
    """Returns (labels, lookup) for use with st.selectbox/st.multiselect:
    `labels` is the display list, `lookup[label]` maps back to the real
    account code via a plain dict — never by parsing the label string apart
    (a real source of bugs when a name itself contains ' - ')."""
    codes = codes if codes is not None else sorted_account_codes()
    labels, lookup = [], {}
    for c in codes:
        info = account_info(c)
        label = f"{TYPE_ICON.get(info['type'], '')} {c} — {info['name']}"
        labels.append(label)
        lookup[label] = c
    return labels, lookup

# ============================================================================
# UTILITY FUNCTIONS — currency formatting
# ============================================================================

def currency_symbol(currency=None) -> str:
    currency = currency or st.session_state.get("currency", "Rp")
    return CURRENCY_META[currency]["symbol"]

def currency_label(base: str, currency=None) -> str:
    """'Debit' -> 'Debit (Rp)' — exactly what shows up in table headers
    and Excel exports."""
    return f"{base} ({currency_symbol(currency)})"

def fmt_money(value, currency=None, dash_on_zero=False) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        value = 0.0
    value = float(value)
    currency = currency or st.session_state.get("currency", "Rp")
    if dash_on_zero and abs(value) < 0.005:
        return "-"
    negative = value < -0.005
    abs_val = abs(value)
    if currency == "Rp":
        s = f"{abs_val:,.0f}".replace(",", ".")
    else:
        s = f"{abs_val:,.2f}"
    s = f"{currency_symbol(currency)} {s}"
    return f"({s})" if negative else s

def money_input_step(currency=None):
    currency = currency or st.session_state.get("currency", "Rp")
    return (1000.0, "%.0f") if currency == "Rp" else (0.01, "%.2f")

def excel_number_format(currency=None) -> str:
    currency = currency or st.session_state.get("currency", "Rp")
    if currency == "Rp":
        return '_-"Rp"* #,##0_-;-"Rp"* #,##0_-;_-"Rp"* "-"_-;_-@_-'
    return '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

# ============================================================================
# VALIDATION
# ============================================================================

def validate_transaction(entries):
    """Returns (is_valid, total_debit, total_credit, message, line_warnings)."""
    total_debit = sum(e["Debit"] for e in entries)
    total_credit = sum(e["Credit"] for e in entries)
    is_balanced = abs(total_debit - total_credit) < 0.01
    has_amount = (total_debit + total_credit) > 0

    warnings = []
    for i, e in enumerate(entries, start=1):
        if e["Debit"] > 0 and e["Credit"] > 0:
            warnings.append(f"Line {i}: has both a Debit and a Credit amount — enter only one per line.")

    if not has_amount:
        return False, total_debit, total_credit, "Enter at least one Debit and one Credit amount.", warnings
    if warnings:
        return False, total_debit, total_credit, "Fix the flagged line(s) before posting.", warnings
    if is_balanced:
        return True, total_debit, total_credit, "Balanced — ready to post.", warnings
    return False, total_debit, total_credit, "Debits and Credits don't match yet.", warnings

# ============================================================================
# CALCULATION ENGINE
# Every report below is built on account_net(), so Trial Balance, Income
# Statement, Statement of Owner's Equity and Balance Sheet can never
# disagree with each other about a balance.
# ============================================================================

def journal_df() -> pd.DataFrame:
    return st.session_state.journal_df

def filter_by_date(df, start=None, end=None):
    out = df
    if start is not None:
        out = out[out["Date"] >= pd.Timestamp(start)]
    if end is not None:
        out = out[out["Date"] <= pd.Timestamp(end)]
    return out

def account_net(df, code, start=None, end=None) -> float:
    """Debit - Credit for one account, optionally bounded by date."""
    sub = df[df["Post_Ref"] == str(code)]
    sub = filter_by_date(sub, start, end)
    if sub.empty:
        return 0.0
    return float(sub["Debit"].sum() - sub["Credit"].sum())

def cumulative_net_income(df, up_to) -> float:
    """Total Revenue minus total Expense from inception through `up_to`."""
    revenue = 0.0
    expense = 0.0
    for code in sorted_account_codes():
        info = account_info(code)
        if info["type"] == "Revenue":
            revenue += -account_net(df, code, None, up_to)
        elif info["type"] == "Expense":
            expense += account_net(df, code, None, up_to)
    return revenue - expense

def compute_trial_balance(df, as_of):
    rows = []
    for code in sorted_account_codes():
        net = account_net(df, code, None, as_of)
        if abs(net) < 0.005:
            continue
        info = account_info(code)
        rows.append({
            "Code": code,
            "Account": info["name"],
            "Type": info["type"],
            "Debit": net if net > 0 else 0.0,
            "Credit": -net if net < 0 else 0.0,
        })
    tb_df = pd.DataFrame(rows, columns=["Code", "Account", "Type", "Debit", "Credit"])
    total_debit = float(tb_df["Debit"].sum()) if not tb_df.empty else 0.0
    total_credit = float(tb_df["Credit"].sum()) if not tb_df.empty else 0.0
    return tb_df, total_debit, total_credit

def compute_income_statement(df, start, end):
    period = filter_by_date(df, start, end)
    revenues, expenses = [], []
    for code in sorted_account_codes():
        info = account_info(code)
        sub = period[period["Post_Ref"] == code]
        if sub.empty:
            continue
        if info["type"] == "Revenue":
            amt = float(sub["Credit"].sum() - sub["Debit"].sum())
            if abs(amt) > 0.005:
                revenues.append((code, info["name"], amt))
        elif info["type"] == "Expense":
            amt = float(sub["Debit"].sum() - sub["Credit"].sum())
            if abs(amt) > 0.005:
                expenses.append((code, info["name"], amt))
    total_revenue = sum(a for *_r, a in revenues)
    total_expenses = sum(a for *_r, a in expenses)
    return {
        "revenues": revenues, "total_revenue": total_revenue,
        "expenses": expenses, "total_expenses": total_expenses,
        "net_income": total_revenue - total_expenses,
    }

def compute_owners_equity(df, start, end, net_income_for_period):
    """Beginning Capital is computed cumulatively (contributions + net
    income - drawings from inception through the day before `start`), not
    just the raw Capital-account balance. That is what makes the Balance
    Sheet tie out exactly no matter what period the user picks, even though
    this system never posts formal closing entries."""
    capital_codes = [c for c in sorted_account_codes() if account_info(c)["subtype"] == "Capital"]
    drawing_codes = [c for c in sorted_account_codes() if account_info(c)["subtype"] == "Drawing"]

    before_start = pd.Timestamp(start) - pd.Timedelta(days=1)

    capital_before = sum(-account_net(df, c, None, before_start) for c in capital_codes)
    drawings_before = sum(account_net(df, c, None, before_start) for c in drawing_codes)
    net_income_before = cumulative_net_income(df, before_start)
    beginning_capital = capital_before + net_income_before - drawings_before

    investment = sum(-account_net(df, c, start, end) for c in capital_codes)
    drawings = sum(account_net(df, c, start, end) for c in drawing_codes)

    ending_capital = beginning_capital + investment + net_income_for_period - drawings

    return {
        "capital_codes": capital_codes, "drawing_codes": drawing_codes,
        "beginning_capital": beginning_capital, "investment": investment,
        "net_income": net_income_for_period, "drawings": drawings,
        "ending_capital": ending_capital,
    }

def compute_balance_sheet(df, as_of, ending_capital):
    asset_groups = {"Current Asset": [], "Fixed Asset": [], "Other Asset": []}
    liability_groups = {"Current Liability": [], "Long-Term Liability": []}

    for code in sorted_account_codes():
        info = account_info(code)
        if info["type"] not in ("Asset", "Liability"):
            continue
        net = account_net(df, code, None, as_of)
        if abs(net) < 0.005:
            continue
        if info["type"] == "Asset":
            asset_groups.setdefault(info["subtype"], []).append(
                (code, info["name"], net, info.get("contra", False))
            )
        else:
            liability_groups.setdefault(info["subtype"], []).append(
                (code, info["name"], -net)
            )

    total_assets = sum(a for items in asset_groups.values() for *_r, a, _c in items)
    total_liabilities = sum(a for items in liability_groups.values() for *_r, a in items)
    total_liab_equity = total_liabilities + ending_capital

    return {
        "asset_groups": asset_groups,
        "liability_groups": liability_groups,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "equity": ending_capital,
        "total_liab_equity": total_liab_equity,
        "is_balanced": abs(total_assets - total_liab_equity) < 0.01,
    }

def compute_account_ledger_rows(df, code):
    """Returns (ledger_rows, ending_balance) for one account: Date,
    Description (contra-account, or 'Sundries' for compound entries),
    Debit, Credit and running Balance. Shared by the on-screen View Ledger
    page and every Excel ledger export so the two can never disagree."""
    acct_txns = df[df["Post_Ref"] == code].sort_values(
        ["Date", "Transaction_ID", "Credit"]
    ).reset_index(drop=True)

    if acct_txns.empty:
        return [], 0.0

    normal_debit = account_normal_balance(code) == "Debit"
    running = 0.0
    rows = []
    for _, r in acct_txns.iterrows():
        running += (r["Debit"] - r["Credit"]) if normal_debit else (r["Credit"] - r["Debit"])
        other_accounts = df[
            (df["Transaction_ID"] == r["Transaction_ID"]) & (df["Post_Ref"] != code)
        ]["Account_Title"].unique()
        description = " / ".join(other_accounts) if len(other_accounts) else (r["Explanation"] or "")
        rows.append({
            "Date": r["Date"], "Description": description,
            "Debit": r["Debit"], "Credit": r["Credit"], "Balance": running,
        })
    return rows, rows[-1]["Balance"]

def accounts_with_activity(df):
    """Account codes that have at least one posting, in Chart-of-Accounts order."""
    if df.empty:
        return []
    posted = set(df["Post_Ref"].unique())
    return [c for c in sorted_account_codes() if c in posted]

# ============================================================================
# EXCEL EXPORT — shared style helpers
# All exports share one visual language: a dark ledger-green title band,
# a light header row, thin ledger-tan gridlines and a double-ruled bottom
# border under final totals (the standard accounting-statement convention).
# ============================================================================

XL_FONT = "Calibri"
XL_TITLE_BG = "1F77B4"
XL_TITLE_FONT = "FFFFFF"
XL_SUBTITLE_FONT = "44545C"
XL_HEADER_BG = "2E75B6"
XL_HEADER_FONT = "FFFFFF"
XL_TOTAL_BG = "DCEAF6"
XL_SUBTOTAL_BG = "EEF4FA"
XL_BORDER_COLOR = "BFD3E4"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    thin = Side(style="thin", color=XL_BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _double_bottom_border():
    thin = Side(style="thin", color=XL_BORDER_COLOR)
    dbl = Side(style="double", color="1C2B33")
    return Border(top=thin, bottom=dbl, left=thin, right=thin)

def _require_openpyxl():
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("The 'openpyxl' package is not installed. Run: pip install openpyxl")

def _write_title_block(ws, title, subtitle, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name=XL_FONT, bold=True, size=15, color=XL_TITLE_FONT)
    c.fill = _fill(XL_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        sc = ws.cell(row=2, column=1, value=subtitle)
        sc.font = Font(name=XL_FONT, italic=True, size=10.5, color=XL_SUBTITLE_FONT)
        sc.fill = _fill("F4F1EA")
        sc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18
        row = 3
    return row + 1  # one blank spacer row before content

def _header_row(ws, row, headers, currency_cols=None):
    currency_cols = currency_cols or set()
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = Font(name=XL_FONT, bold=True, size=10.5, color=XL_HEADER_FONT)
        c.fill = _fill(XL_HEADER_BG)
        c.alignment = Alignment(
            horizontal="right" if idx in currency_cols else "left",
            vertical="center", indent=0 if idx in currency_cols else 1,
        )
        c.border = _thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1

# -- Two-column "Label | Amount" statement writer (Income Statement, --------
# -- Statement of Owner's Equity, Balance Sheet) ----------------------------

def _write_statement_rows(ws, row, blocks, currency, label_col=1, amount_col=2, num_cols=2):
    num_fmt = excel_number_format(currency)
    for block in blocks:
        kind = block[0]
        if kind == "spacer":
            row += 1
            continue
        if kind == "header":
            text = block[1]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(name=XL_FONT, bold=True, size=12, color=XL_HEADER_FONT)
            c.fill = _fill(XL_HEADER_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1
        elif kind == "subheader":
            text = block[1]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(name=XL_FONT, bold=True, italic=True, size=10.5, color=XL_TITLE_BG)
            c.fill = _fill(XL_SUBTOTAL_BG)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row += 1
        elif kind == "line":
            label, amount = block[1], block[2]
            indent = block[3] if len(block) > 3 else 2
            lc = ws.cell(row=row, column=label_col, value=label)
            lc.font = Font(name=XL_FONT, size=10.5)
            lc.alignment = Alignment(horizontal="left", indent=indent)
            ac = ws.cell(row=row, column=amount_col, value=round(float(amount), 2))
            ac.font = Font(name=XL_FONT, size=10.5)
            ac.number_format = num_fmt
            ac.alignment = Alignment(horizontal="right")
            row += 1
        elif kind == "subtotal":
            label, amount = block[1], block[2]
            border = Border(top=Side(style="thin", color=XL_BORDER_COLOR))
            lc = ws.cell(row=row, column=label_col, value=label)
            lc.font = Font(name=XL_FONT, bold=True, size=10.5)
            lc.alignment = Alignment(horizontal="left", indent=1)
            lc.border = border
            ac = ws.cell(row=row, column=amount_col, value=round(float(amount), 2))
            ac.font = Font(name=XL_FONT, bold=True, size=10.5)
            ac.number_format = num_fmt
            ac.alignment = Alignment(horizontal="right")
            ac.border = border
            row += 1
        elif kind == "total":
            label, amount = block[1], block[2]
            fill = _fill(XL_TOTAL_BG)
            border = _double_bottom_border()
            lc = ws.cell(row=row, column=label_col, value=label)
            lc.font = Font(name=XL_FONT, bold=True, size=11)
            lc.fill = fill
            lc.border = border
            lc.alignment = Alignment(horizontal="left", indent=1)
            ac = ws.cell(row=row, column=amount_col, value=round(float(amount), 2))
            ac.font = Font(name=XL_FONT, bold=True, size=11)
            ac.number_format = num_fmt
            ac.fill = fill
            ac.border = border
            ac.alignment = Alignment(horizontal="right")
            row += 1
    return row

def _finalize(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ============================================================================
# EXCEL EXPORT — General Journal
# ============================================================================

def generate_excel_journal(txn_groups, currency, workbook=None):
    """txn_groups: list of (transaction_id, date, [line dicts]) already
    sorted chronologically, each line dict having Account_Title, Post_Ref,
    Explanation, Debit, Credit.

    Pass `workbook` to append this report as a sheet inside a larger,
    already-open workbook (used by the full-cycle export); leave it None
    to get a standalone .xlsx buffer back (used by the per-page button)."""
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "General Journal"
    else:
        ws = wb.create_sheet("General Journal")

    widths = [13, 34, 10, 30, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = _write_title_block(ws, "General Journal", "For All Recorded Transactions", 6)
    headers = ["Date", "Account Title", "Post Ref.", "Explanation",
               currency_label("Debit", currency), currency_label("Credit", currency)]
    row = _header_row(ws, row, headers, currency_cols={5, 6})

    num_fmt = excel_number_format(currency)
    first_data_row = row
    thin = _thin_border()

    for txn_id, txn_date, lines in txn_groups:
        for i, line in enumerate(lines):
            is_credit_line = line["Credit"] > 0
            title = ("    " + line["Account_Title"]) if is_credit_line else line["Account_Title"]
            date_val = txn_date.date() if i == 0 else None

            c1 = ws.cell(row=row, column=1, value=date_val)
            if date_val is not None:
                c1.number_format = "dd-mmm-yyyy"
            c2 = ws.cell(row=row, column=2, value=title)
            c3 = ws.cell(row=row, column=3, value=str(line["Post_Ref"]))
            c4 = ws.cell(row=row, column=4, value=line.get("Explanation", "") if i == 0 else "")
            c5 = ws.cell(row=row, column=5, value=round(float(line["Debit"]), 2) if line["Debit"] else None)
            c6 = ws.cell(row=row, column=6, value=round(float(line["Credit"]), 2) if line["Credit"] else None)

            for c in (c1, c2, c3, c4, c5, c6):
                c.border = thin
                c.font = Font(name=XL_FONT, size=10.5)
            c1.alignment = Alignment(horizontal="center")
            c3.alignment = Alignment(horizontal="center")
            c2.alignment = Alignment(horizontal="left", indent=2 if is_credit_line else 1)
            c5.number_format = num_fmt
            c6.number_format = num_fmt
            row += 1
        row += 1  # blank spacer row between transactions

    last_data_row = row - 1
    total_row = row
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(name=XL_FONT, bold=True)
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    tc5 = ws.cell(row=total_row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
    tc6 = ws.cell(row=total_row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
    for c in (ws.cell(row=total_row, column=4), tc5, tc6):
        c.fill = _fill(XL_TOTAL_BG)
        c.border = _double_bottom_border()
        c.font = Font(name=XL_FONT, bold=True)
    tc5.number_format = num_fmt
    tc6.number_format = num_fmt

    ws.freeze_panes = f"A{first_data_row}"
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — General Ledger (every account with activity, one sheet)
# ============================================================================

def generate_excel_all_ledgers(df, currency, workbook=None):
    """Writes ONE 'General Ledger' sheet containing every account that has
    postings, each as its own titled block with a running balance — this
    mirrors the on-screen View Ledger page (which also shows every account,
    not just one picked from a dropdown)."""
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "General Ledger"
    else:
        ws = wb.create_sheet("General Ledger")

    widths = [13, 42, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = _write_title_block(ws, "General Ledger", "Every Account With Posted Activity", 5)
    num_fmt = excel_number_format(currency)
    thin = _thin_border()

    codes = accounts_with_activity(df)

    if not codes:
        ws.cell(row=row, column=1, value="No postings yet.").font = Font(name=XL_FONT, italic=True)
        if standalone:
            return _finalize(wb)
        return

    for code in codes:
        info = account_info(code)
        ledger_rows, ending_balance = compute_account_ledger_rows(df, code)

        # -- Account title band -------------------------------------------
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        hc = ws.cell(
            row=row, column=1,
            value=f"{code} — {info['name']}  ({info['type']} · {info['subtype']})",
        )
        hc.font = Font(name=XL_FONT, bold=True, size=11, color=XL_HEADER_FONT)
        hc.fill = _fill(XL_HEADER_BG)
        hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        # -- Column headers for this account's block -----------------------
        headers = ["Date", "Description", currency_label("Debit", currency),
                   currency_label("Credit", currency), currency_label("Balance", currency)]
        for idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=idx, value=h)
            c.font = Font(name=XL_FONT, bold=True, italic=True, size=10)
            c.fill = _fill(XL_SUBTOTAL_BG)
            c.alignment = Alignment(
                horizontal="right" if idx >= 3 else "left",
                vertical="center", indent=0 if idx >= 3 else 1,
            )
            c.border = thin
        row += 1

        first_data_row = row
        for r in ledger_rows:
            c1 = ws.cell(row=row, column=1, value=r["Date"].date() if hasattr(r["Date"], "date") else r["Date"])
            c1.number_format = "dd-mmm-yyyy"
            c2 = ws.cell(row=row, column=2, value=r["Description"])
            c3 = ws.cell(row=row, column=3, value=round(float(r["Debit"]), 2) if r["Debit"] else None)
            c4 = ws.cell(row=row, column=4, value=round(float(r["Credit"]), 2) if r["Credit"] else None)
            c5 = ws.cell(row=row, column=5, value=round(float(r["Balance"]), 2))
            for c in (c1, c2, c3, c4, c5):
                c.border = thin
                c.font = Font(name=XL_FONT, size=10)
            c1.alignment = Alignment(horizontal="center")
            c2.alignment = Alignment(horizontal="left", indent=1)
            for c in (c3, c4, c5):
                c.number_format = num_fmt
            row += 1
        last_data_row = row - 1

        # -- Ending balance row for this account -----------------------------
        ws.cell(row=row, column=2, value="Ending Balance").font = Font(name=XL_FONT, bold=True, size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        tc3 = ws.cell(row=row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})" if last_data_row >= first_data_row else 0)
        tc4 = ws.cell(row=row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})" if last_data_row >= first_data_row else 0)
        tc5 = ws.cell(row=row, column=5, value=round(float(ending_balance), 2))
        for c in (ws.cell(row=row, column=2), tc3, tc4, tc5):
            c.fill = _fill(XL_TOTAL_BG)
            c.border = _double_bottom_border()
            c.font = Font(name=XL_FONT, bold=True, size=10)
        tc3.number_format = num_fmt
        tc4.number_format = num_fmt
        tc5.number_format = num_fmt
        row += 2  # spacer row between accounts

    ws.freeze_panes = "A1"
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — Trial Balance
# ============================================================================

def generate_excel_trial_balance(tb_df, as_of, total_debit, total_credit, currency, workbook=None):
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "Trial Balance"
    else:
        ws = wb.create_sheet("Trial Balance")

    widths = [12, 36, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    date_str = as_of.strftime("%d %B %Y") if hasattr(as_of, "strftime") else str(as_of)
    row = _write_title_block(ws, "Trial Balance", f"As of {date_str}", 4)
    headers = ["Code", "Account", currency_label("Debit", currency), currency_label("Credit", currency)]
    row = _header_row(ws, row, headers, currency_cols={3, 4})

    num_fmt = excel_number_format(currency)
    thin = _thin_border()
    first_data_row = row
    for _, r in tb_df.iterrows():
        c1 = ws.cell(row=row, column=1, value=str(r["Code"]))
        c2 = ws.cell(row=row, column=2, value=r["Account"])
        c3 = ws.cell(row=row, column=3, value=round(float(r["Debit"]), 2) if r["Debit"] else None)
        c4 = ws.cell(row=row, column=4, value=round(float(r["Credit"]), 2) if r["Credit"] else None)
        for c in (c1, c2, c3, c4):
            c.border = thin
            c.font = Font(name=XL_FONT, size=10.5)
        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="left", indent=1)
        c3.number_format = num_fmt
        c4.number_format = num_fmt
        row += 1
    last_data_row = row - 1

    total_row = row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(name=XL_FONT, bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    tc3 = ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})" if last_data_row >= first_data_row else 0)
    tc4 = ws.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})" if last_data_row >= first_data_row else 0)
    for c in (ws.cell(row=total_row, column=2), tc3, tc4):
        c.fill = _fill(XL_TOTAL_BG)
        c.border = _double_bottom_border()
        c.font = Font(name=XL_FONT, bold=True)
    tc3.number_format = num_fmt
    tc4.number_format = num_fmt

    note_row = total_row + 2
    is_balanced = abs(total_debit - total_credit) < 0.01
    note = "✓ Trial Balance is in balance." if is_balanced else "⚠ Trial Balance is NOT in balance — check postings."
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    nc = ws.cell(row=note_row, column=1, value=note)
    nc.font = Font(name=XL_FONT, italic=True, bold=not is_balanced,
                    color="145A8A" if is_balanced else "B00020")

    ws.freeze_panes = f"A{first_data_row}"
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — Income Statement
# ============================================================================

def generate_excel_income_statement(stmt, start, end, currency, workbook=None):
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "Income Statement"
    else:
        ws = wb.create_sheet("Income Statement")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22

    subtitle = f"For the Period {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    row = _write_title_block(ws, "Income Statement", subtitle, 2)

    blocks = [("header", "REVENUES")]
    for _, nm, amt in stmt["revenues"]:
        blocks.append(("line", nm, amt))
    if not stmt["revenues"]:
        blocks.append(("line", "(no revenue recorded)", 0))
    blocks.append(("subtotal", "Total Revenue", stmt["total_revenue"]))
    blocks.append(("spacer",))
    blocks.append(("header", "EXPENSES"))
    for _, nm, amt in stmt["expenses"]:
        blocks.append(("line", nm, amt))
    if not stmt["expenses"]:
        blocks.append(("line", "(no expenses recorded)", 0))
    blocks.append(("subtotal", "Total Expenses", stmt["total_expenses"]))
    blocks.append(("spacer",))
    label = "NET INCOME" if stmt["net_income"] >= 0 else "NET LOSS"
    blocks.append(("total", label, stmt["net_income"]))

    _write_statement_rows(ws, row, blocks, currency)
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — Statement of Owner's Equity
# ============================================================================

def generate_excel_owners_equity(stmt, start, end, currency, workbook=None):
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "Owner's Equity"
    else:
        ws = wb.create_sheet("Owner's Equity")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    subtitle = f"For the Period {start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"
    row = _write_title_block(ws, "Statement of Owner's Equity", subtitle, 2)

    blocks = [("line", "Owner's Capital, Beginning", stmt["beginning_capital"], 1)]
    if abs(stmt["investment"]) > 0.005:
        blocks.append(("line", "Add: Owner's Investment", stmt["investment"], 1))
    if stmt["net_income"] >= 0:
        blocks.append(("line", "Add: Net Income", stmt["net_income"], 1))
    else:
        blocks.append(("line", "Less: Net Loss", stmt["net_income"], 1))
    if abs(stmt["drawings"]) > 0.005:
        blocks.append(("line", "Less: Owner's Drawings", -stmt["drawings"], 1))
    blocks.append(("spacer",))
    blocks.append(("total", "Owner's Capital, Ending", stmt["ending_capital"]))

    _write_statement_rows(ws, row, blocks, currency)
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — Balance Sheet
# ============================================================================

def generate_excel_balance_sheet(stmt, as_of, currency, workbook=None):
    _require_openpyxl()
    standalone = workbook is None
    wb = workbook if workbook is not None else Workbook()
    if standalone:
        ws = wb.active
        ws.title = "Balance Sheet"
    else:
        ws = wb.create_sheet("Balance Sheet")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 22

    date_str = as_of.strftime("%d %B %Y") if hasattr(as_of, "strftime") else str(as_of)
    row = _write_title_block(ws, "Balance Sheet", f"As of {date_str}", 2)

    blocks = [("header", "ASSETS")]
    group_order = ["Current Asset", "Fixed Asset", "Other Asset"]
    group_titles = {"Current Asset": "Current Assets", "Fixed Asset": "Fixed Assets", "Other Asset": "Other Assets"}
    for g in group_order:
        items = stmt["asset_groups"].get(g, [])
        if not items:
            continue
        blocks.append(("subheader", group_titles[g]))
        subtotal = 0.0
        for _code, nm, amount, is_contra in items:
            label = f"Less: {nm}" if is_contra else nm
            blocks.append(("line", label, amount))
            subtotal += amount
        subtotal_label = "Net Fixed Assets" if g == "Fixed Asset" else f"Total {group_titles[g]}"
        blocks.append(("subtotal", subtotal_label, subtotal))
        blocks.append(("spacer",))
    blocks.append(("total", "TOTAL ASSETS", stmt["total_assets"]))
    blocks.append(("spacer",))
    blocks.append(("spacer",))

    blocks.append(("header", "LIABILITIES"))
    l_order = ["Current Liability", "Long-Term Liability"]
    l_titles = {"Current Liability": "Current Liabilities", "Long-Term Liability": "Long-Term Liabilities"}
    any_liability = False
    for g in l_order:
        items = stmt["liability_groups"].get(g, [])
        if not items:
            continue
        any_liability = True
        blocks.append(("subheader", l_titles[g]))
        subtotal = 0.0
        for _code, nm, amount in items:
            blocks.append(("line", nm, amount))
            subtotal += amount
        blocks.append(("subtotal", f"Total {l_titles[g]}", subtotal))
        blocks.append(("spacer",))
    if not any_liability:
        blocks.append(("line", "(no liabilities recorded)", 0))
        blocks.append(("spacer",))
    blocks.append(("subtotal", "Total Liabilities", stmt["total_liabilities"]))
    blocks.append(("spacer",))

    blocks.append(("header", "OWNER'S EQUITY"))
    blocks.append(("line", "Owner's Capital, Ending", stmt["equity"], 2))
    blocks.append(("subtotal", "Total Owner's Equity", stmt["equity"]))
    blocks.append(("spacer",))
    blocks.append(("total", "TOTAL LIABILITIES & OWNER'S EQUITY", stmt["total_liab_equity"]))

    _write_statement_rows(ws, row, blocks, currency)
    if standalone:
        return _finalize(wb)

# ============================================================================
# EXCEL EXPORT — Full Accounting Cycle (all reports, one workbook)
# ============================================================================

def generate_excel_full_cycle(df, currency):
    """One workbook covering the complete cycle, sheet by sheet, in order:
    General Journal -> General Ledger -> Trial Balance -> Income Statement ->
    Statement of Owner's Equity -> Balance Sheet. Reuses the exact same
    calculation and export functions as the individual per-page buttons, so
    the numbers can never drift out of sync with what's shown on screen."""
    _require_openpyxl()
    wb = Workbook()
    wb.remove(wb.active)  # start empty; each report below adds its own sheet

    inception = df["Date"].min()
    latest = df["Date"].max()

    # 1. General Journal — every transaction, chronologically
    groups = []
    for txn_id, grp in df.groupby("Transaction_ID", sort=False):
        grp = grp.sort_values("Credit")
        groups.append((txn_id, grp["Date"].iloc[0], grp.to_dict("records")))
    groups.sort(key=lambda g: g[1])
    generate_excel_journal(groups, currency, workbook=wb)

    # 2. General Ledger — every account with postings
    generate_excel_all_ledgers(df, currency, workbook=wb)

    # 3. Trial Balance — as of the latest transaction date
    tb_df, total_debit, total_credit = compute_trial_balance(df, latest)
    generate_excel_trial_balance(tb_df, latest, total_debit, total_credit, currency, workbook=wb)

    # 4-6. Financial statements — full period, inception through latest date
    income = compute_income_statement(df, inception, latest)
    generate_excel_income_statement(income, inception, latest, currency, workbook=wb)

    equity = compute_owners_equity(df, inception, latest, income["net_income"])
    generate_excel_owners_equity(equity, inception, latest, currency, workbook=wb)

    bs = compute_balance_sheet(df, latest, equity["ending_capital"])
    generate_excel_balance_sheet(bs, latest, currency, workbook=wb)

    return _finalize(wb)

# ============================================================================
# SIDEBAR / NAVIGATION
# ============================================================================

NAV_GROUPS = [
    ("WORKFLOW", [
        ("Input Journal", "✍️"),
        ("View Journal", "📖"),
        ("View Ledger", "📚"),
    ]),
    ("REPORTS", [
        ("Trial Balance", "⚖️"),
        ("Financial Statements", "📊"),
    ]),
    ("EXPORT", [
        ("Export Full Cycle", "📦"),
    ]),
    ("SETUP", [
        ("Chart of Accounts", "🗂️"),
    ]),
]

def render_sidebar():
    with st.sidebar:
        st.markdown("## 💰 Accounting System")
        st.caption("General Journal → Ledger → Financial Statements")
        st.markdown("---")

        for group_name, items in NAV_GROUPS:
            st.markdown(f'<div class="nav-eyebrow">{group_name}</div>', unsafe_allow_html=True)
            for label, icon in items:
                is_active = st.session_state.nav_page == label
                if st.button(
                    f"{icon}  {label}",
                    key=f"nav_{label}",
                    width="stretch",
                    type="primary" if is_active else "secondary",
                ):
                    st.session_state.nav_page = label
                    st.rerun()

        st.markdown("---")
        st.markdown('<div class="nav-eyebrow">CURRENCY</div>', unsafe_allow_html=True)
        cur_choice = st.selectbox(
            "Report currency", options=["Rp", "$"],
            index=["Rp", "$"].index(st.session_state.currency),
            key="currency_selector", label_visibility="collapsed",
            format_func=lambda c: "Rupiah (Rp)" if c == "Rp" else "US Dollar ($)",
        )
        if cur_choice != st.session_state.currency:
            st.session_state.currency = cur_choice
            st.session_state.app_settings["currency"] = cur_choice
            data_mgr.save_settings(st.session_state.app_settings)
            st.rerun()
        st.caption("Changes labels & formatting only — amounts aren't converted.")

        st.markdown("---")
        df = journal_df()
        n_txn = df["Transaction_ID"].nunique() if not df.empty else 0
        n_accts = len(coa())
        st.caption(f"📄 {n_txn} transaction(s) · 🗂️ {n_accts} account(s)")
        st.caption(f"💾 Data folder: `{data_mgr.data_dir.name}/`")

# ============================================================================
# PAGE — Input Journal
# ============================================================================

def page_input_journal():
    st.title("✍️ Input Journal")
    st.caption("Record a balanced double-entry transaction (total Debit must equal total Credit).")

    if len(coa()) == 0:
        st.warning("No accounts exist yet. Add some in **Chart of Accounts** first.")
        return

    fk = st.session_state.entry_form_key  # bumped after Post/Clear to reset widget values
    acct_labels, acct_lookup = account_select_options()

    col_date, col_num = st.columns([2, 1])
    with col_date:
        txn_date = st.date_input("Transaction Date", value=date.today(), key="txn_date")
    with col_num:
        num_entries = st.number_input("Number of Lines", min_value=2, max_value=12, value=2, step=1, key="num_entries")

    step, fmt_str = money_input_step()

    col1, col2 = st.columns([2, 1])

    entries = []
    with col1:
        st.markdown("#### Transaction Lines")
        for i in range(int(num_entries)):
            st.markdown(f'<div class="ledger-eyebrow">LINE {i + 1}</div>', unsafe_allow_html=True)
            c_acct, c_deb, c_cred = st.columns([2.2, 1, 1])
            with c_acct:
                sel_label = st.selectbox(
                    "Account", options=acct_labels,
                    key=f"acct_{fk}_{i}", label_visibility="collapsed" if i > 0 else "visible",
                )
                sel_code = acct_lookup[sel_label]
            with c_deb:
                debit_val = st.number_input(
                    currency_label("Debit"), min_value=0.0, step=step, format=fmt_str,
                    key=f"debit_{fk}_{i}", label_visibility="collapsed" if i > 0 else "visible",
                )
            with c_cred:
                credit_val = st.number_input(
                    currency_label("Credit"), min_value=0.0, step=step, format=fmt_str,
                    key=f"credit_{fk}_{i}", label_visibility="collapsed" if i > 0 else "visible",
                )
            explanation = st.text_input(
                "Explanation (optional)", key=f"expl_{fk}_{i}",
                label_visibility="collapsed" if i > 0 else "visible",
                placeholder="Explanation (optional)",
            )
            entries.append({
                "Account": sel_code, "Debit": float(debit_val), "Credit": float(credit_val),
                "Explanation": explanation,
            })
            st.markdown("")

    total_debit = sum(e["Debit"] for e in entries)
    total_credit = sum(e["Credit"] for e in entries)
    is_valid, _td, _tc, message, line_warnings = validate_transaction(entries)

    with col2:
        st.markdown("#### Balance Check")
        st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
        st.metric("Total Debit", fmt_money(total_debit))
        st.metric("Total Credit", fmt_money(total_credit))
        diff = total_debit - total_credit
        st.metric("Difference", fmt_money(diff))
        st.markdown("</div>", unsafe_allow_html=True)

        for w in line_warnings:
            st.warning(w)
        if is_valid:
            st.success(message)
        elif not line_warnings:
            st.info(message)

        st.markdown("")
        bcol1, bcol2 = st.columns(2)
        post_clicked = bcol1.button("✅ Post Transaction", type="primary", width="stretch", disabled=not is_valid)
        clear_clicked = bcol2.button("🧹 Clear Form", width="stretch")

    if clear_clicked:
        st.session_state.entry_form_key += 1
        st.rerun()

    if post_clicked and is_valid:
        txn_id = str(uuid.uuid4())[:8]
        new_rows = []
        for e in entries:
            if e["Debit"] == 0 and e["Credit"] == 0:
                continue
            new_rows.append({
                "Date": pd.Timestamp(txn_date),
                "Transaction_ID": txn_id,
                "Line_ID": str(uuid.uuid4()),
                "Account_Title": account_name(e["Account"]),
                "Post_Ref": str(e["Account"]),
                "Explanation": e["Explanation"],
                "Debit": e["Debit"],
                "Credit": e["Credit"],
            })
        new_df = pd.DataFrame(new_rows, columns=JOURNAL_COLUMNS)
        st.session_state.journal_df = pd.concat([st.session_state.journal_df, new_df], ignore_index=True)
        data_mgr.save_journal(st.session_state.journal_df)
        st.session_state.entry_form_key += 1
        st.success(f"Transaction posted ({len(new_rows)} line(s)).")
        st.balloons()
        st.rerun()

    st.markdown("---")
    with st.expander("📎 Chart of Accounts reference"):
        for t in ACCOUNT_TYPES:
            codes_of_type = [c for c in sorted_account_codes() if account_info(c)["type"] == t]
            if not codes_of_type:
                continue
            st.markdown(f"**{TYPE_ICON[t]} {t}**")
            st.caption(" · ".join(f"{c} — {account_name(c)}" for c in codes_of_type))

# ============================================================================
# PAGE — View Journal
# ============================================================================

def page_view_journal():
    st.title("📖 View Journal")
    df = journal_df()

    if df.empty:
        st.info("No transactions recorded yet.")
        return

    with st.expander("🔎 Filters", expanded=False):
        fc1, fc2 = st.columns(2)
        with fc1:
            date_range = st.date_input(
                "Date range", value=(df["Date"].min().date(), df["Date"].max().date()),
                key="vj_date_range",
            )
        with fc2:
            acct_labels, acct_lookup = account_select_options()
            acct_filter_labels = st.multiselect("Filter by account", options=acct_labels, key="vj_acct_filter")
            acct_filter = [acct_lookup[lbl] for lbl in acct_filter_labels]

    filtered = df.copy()
    if isinstance(date_range, tuple) and len(date_range) == 2:
        filtered = filter_by_date(filtered, date_range[0], date_range[1])
    if acct_filter:
        filtered = filtered[filtered["Post_Ref"].isin(acct_filter)]

    filtered = filtered.sort_values(["Date", "Transaction_ID", "Credit"]).reset_index(drop=True)

    st.markdown(f"**{filtered['Transaction_ID'].nunique()}** transaction(s), **{len(filtered)}** line(s)")

    editor_df = filtered[["Date", "Account_Title", "Post_Ref", "Explanation", "Debit", "Credit", "Transaction_ID"]].copy()
    edited = st.data_editor(
        editor_df,
        hide_index=True,
        width="stretch",
        num_rows="fixed",
        key="journal_editor",
        column_config={
            "Date": st.column_config.DateColumn("Date", format="DD-MMM-YYYY"),
            "Account_Title": st.column_config.TextColumn("Account", disabled=True),
            "Post_Ref": st.column_config.TextColumn("Post Ref.", disabled=True),
            "Explanation": st.column_config.TextColumn("Explanation"),
            "Debit": st.column_config.NumberColumn(currency_label("Debit"), disabled=True, format="%.2f"),
            "Credit": st.column_config.NumberColumn(currency_label("Credit"), disabled=True, format="%.2f"),
            "Transaction_ID": st.column_config.TextColumn("Txn ID", disabled=True),
        },
    )
    st.caption("Only **Date** and **Explanation** are editable here, to protect double-entry balance. "
               "To change an amount, delete the transaction below and re-enter it.")

    if st.button("💾 Save Changes", type="primary"):
        if len(edited) == len(filtered):
            try:
                edited = edited.copy()
                edited["Date"] = pd.to_datetime(edited["Date"])
                edited["Line_ID"] = filtered["Line_ID"].values  # safe: fixed row count, stable order
                master = st.session_state.journal_df.set_index("Line_ID")
                for _, r in edited.iterrows():
                    lid = r["Line_ID"]
                    if lid in master.index:
                        master.at[lid, "Date"] = r["Date"]
                        master.at[lid, "Explanation"] = r["Explanation"]
                master = master.reset_index()[JOURNAL_COLUMNS]
                master = master.sort_values(["Date", "Transaction_ID", "Credit"]).reset_index(drop=True)
                st.session_state.journal_df = master
                data_mgr.save_journal(master)
                st.success("Changes saved.")
                st.rerun()
            except Exception as exc:
                st.error(f"Could not save changes: {exc}")
        else:
            st.warning("Nothing to save.")

    st.markdown("---")
    excel_col1, excel_col2 = st.columns([1, 3])
    with excel_col1:
        if st.button("📥 Export to Excel"):
            try:
                groups = []
                for txn_id, grp in filtered.groupby("Transaction_ID", sort=False):
                    grp = grp.sort_values("Credit")
                    groups.append((txn_id, grp["Date"].iloc[0], grp.to_dict("records")))
                groups.sort(key=lambda g: g[1])
                buf = generate_excel_journal(groups, st.session_state.currency)
                st.download_button(
                    "⬇️ Download general_journal.xlsx", data=buf,
                    file_name="general_journal.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as exc:
                st.error(str(exc))

    st.markdown("---")
    st.markdown("#### 🗑️ Delete a Transaction")
    st.caption("Deletes both sides of the transaction together, to keep the ledger balanced.")

    txn_options, txn_map = [], {}
    for tid, grp in df.groupby("Transaction_ID", sort=False):
        d = grp["Date"].iloc[0]
        accts = " / ".join(grp["Account_Title"].unique()[:3])
        amt = grp["Debit"].sum()
        label = f"{d:%d-%b-%Y} · {accts} · {fmt_money(amt)}"
        txn_options.append(label)
        txn_map[label] = tid

    if txn_options:
        chosen_label = st.selectbox("Select a transaction", options=sorted(txn_options, reverse=True), key="del_txn_select")
        chosen_tid = txn_map[chosen_label]

        if st.session_state.confirm_delete_txn == chosen_tid:
            st.warning("Delete this transaction permanently? This cannot be undone.")
            cc1, cc2 = st.columns(2)
            if cc1.button("✅ Yes, delete it", type="primary"):
                st.session_state.journal_df = df[df["Transaction_ID"] != chosen_tid].reset_index(drop=True)
                data_mgr.save_journal(st.session_state.journal_df)
                st.session_state.confirm_delete_txn = None
                st.success("Transaction deleted.")
                st.rerun()
            if cc2.button("Cancel"):
                st.session_state.confirm_delete_txn = None
                st.rerun()
        else:
            if st.button("🗑️ Delete Selected Transaction"):
                st.session_state.confirm_delete_txn = chosen_tid
                st.rerun()

    st.markdown("")
    if st.session_state.confirm_delete_all:
        st.error("Clear **all** journal entries permanently? This cannot be undone.")
        cc1, cc2 = st.columns(2)
        if cc1.button("✅ Yes, clear everything", type="primary"):
            st.session_state.journal_df = pd.DataFrame(columns=JOURNAL_COLUMNS)
            data_mgr.save_journal(st.session_state.journal_df)
            st.session_state.confirm_delete_all = False
            st.success("All journal entries cleared.")
            st.rerun()
        if cc2.button("Cancel", key="cancel_clear_all"):
            st.session_state.confirm_delete_all = False
            st.rerun()
    else:
        if st.button("🧨 Clear All Entries"):
            st.session_state.confirm_delete_all = True
            st.rerun()

# ============================================================================
# PAGE — View Ledger
# ============================================================================

def page_view_ledger():
    st.title("📚 View Ledger")
    st.caption("Every account that has posted activity — shown in full, no need to pick one at a time.")
    df = journal_df()

    if df.empty:
        st.info("No transactions recorded yet.")
        return

    codes = accounts_with_activity(df)

    if not codes:
        st.info("No postings yet.")
        return

    st.metric("Accounts With Activity", len(codes))
    st.markdown("---")

    for code in codes:
        info = account_info(code)
        ledger_rows, ending_balance = compute_account_ledger_rows(df, code)

        st.markdown(f"### {TYPE_ICON.get(info['type'], '')} {code} — {info['name']}")
        hcol1, hcol2, hcol3 = st.columns(3)
        hcol1.metric("Type", f"{info['type']} · {info['subtype']}")
        hcol2.metric("Normal Balance", account_normal_balance(code))
        hcol3.metric("Ending Balance", fmt_money(ending_balance))

        ledger_df = pd.DataFrame(ledger_rows)
        show = ledger_df.copy()
        show["Date"] = show["Date"].dt.strftime("%d %b %Y")
        for c in ["Debit", "Credit", "Balance"]:
            show[c] = ledger_df[c].apply(lambda v: fmt_money(v, dash_on_zero=(c != "Balance")))
        show.columns = ["Date", "Description", currency_label("Debit"), currency_label("Credit"), currency_label("Balance")]
        st.dataframe(show, hide_index=True, width="stretch")
        st.markdown("---")

    if st.button("📥 Export All Ledgers to Excel", type="primary"):
        try:
            buf = generate_excel_all_ledgers(df, st.session_state.currency)
            st.download_button(
                "⬇️ Download general_ledger.xlsx", data=buf,
                file_name="general_ledger_all_accounts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(str(exc))

# ============================================================================
# PAGE — Trial Balance
# ============================================================================

def page_trial_balance():
    st.title("⚖️ Trial Balance")
    df = journal_df()

    if df.empty:
        st.info("No transactions recorded yet.")
        return

    as_of = st.date_input("As of date", value=df["Date"].max().date(), key="tb_as_of")
    tb_df, total_debit, total_credit = compute_trial_balance(df, as_of)

    if tb_df.empty:
        st.info("No account activity as of this date.")
        return

    show = tb_df.copy()
    show["Debit"] = show["Debit"].apply(lambda v: fmt_money(v, dash_on_zero=True))
    show["Credit"] = show["Credit"].apply(lambda v: fmt_money(v, dash_on_zero=True))
    show = show[["Code", "Account", "Type", "Debit", "Credit"]]
    show.columns = ["Code", "Account", "Type", currency_label("Debit"), currency_label("Credit")]
    st.dataframe(show, hide_index=True, width="stretch")

    c1, c2, c3 = st.columns(3)
    c1.metric("Total Debit", fmt_money(total_debit))
    c2.metric("Total Credit", fmt_money(total_credit))
    is_balanced = abs(total_debit - total_credit) < 0.01
    c3.markdown(
        f'<div style="padding-top:1.6rem;">{"✅ <span class=\'status-balanced\'>Balanced</span>" if is_balanced else "⚠️ <span class=\'status-unbalanced\'>Not balanced</span>"}</div>',
        unsafe_allow_html=True,
    )

    if st.button("📥 Export Trial Balance to Excel"):
        try:
            buf = generate_excel_trial_balance(tb_df, as_of, total_debit, total_credit, st.session_state.currency)
            st.download_button(
                "⬇️ Download trial_balance.xlsx", data=buf,
                file_name="trial_balance.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(str(exc))

# ============================================================================
# PAGE — Financial Statements
# ============================================================================

def _statement_line(label, amount, bold=False, indent=False):
    weight = "font-weight:700;" if bold else ""
    pad = "padding-left:1.2rem;" if indent else ""
    st.markdown(
        f'<div class="statement-line" style="{weight}{pad}"><span>{label}</span><span>{fmt_money(amount)}</span></div>',
        unsafe_allow_html=True,
    )

def _statement_subtotal(label, amount):
    st.markdown(
        f'<div class="statement-subtotal"><span>{label}</span><span>{fmt_money(amount)}</span></div>',
        unsafe_allow_html=True,
    )

def _statement_total(label, amount):
    st.markdown(
        f'<div class="statement-total"><span>{label}</span><span>{fmt_money(amount)}</span></div>',
        unsafe_allow_html=True,
    )

def page_financial_statements():
    st.title("📊 Financial Statements")
    df = journal_df()

    if df.empty:
        st.info("No transactions recorded yet.")
        return

    st.caption("Choose the reporting period. For statements to tie out correctly, Period Start "
               "is usually the very first day of business — the app rolls forward every prior "
               "period's net income automatically either way, so the Balance Sheet always balances.")
    d1, d2 = st.columns(2)
    with d1:
        start = st.date_input("Period Start", value=df["Date"].min().date(), key="fs_start")
    with d2:
        end = st.date_input("Period End", value=df["Date"].max().date(), key="fs_end")

    if start > end:
        st.error("Period Start must be on or before Period End.")
        return

    income = compute_income_statement(df, start, end)
    equity = compute_owners_equity(df, start, end, income["net_income"])
    bs = compute_balance_sheet(df, end, equity["ending_capital"])

    tab1, tab2, tab3 = st.tabs(["📊 Income Statement", "👤 Statement of Owner's Equity", "🏛️ Balance Sheet"])

    with tab1:
        st.markdown(f"##### For the Period {start:%d %b %Y} to {end:%d %b %Y}")
        st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
        st.markdown("**Revenues**")
        if income["revenues"]:
            for _, nm, amt in income["revenues"]:
                _statement_line(nm, amt, indent=True)
        else:
            st.caption("(no revenue recorded)")
        _statement_subtotal("Total Revenue", income["total_revenue"])
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("**Expenses**")
        if income["expenses"]:
            for _, nm, amt in income["expenses"]:
                _statement_line(nm, amt, indent=True)
        else:
            st.caption("(no expenses recorded)")
        _statement_subtotal("Total Expenses", income["total_expenses"])
        st.markdown("<br>", unsafe_allow_html=True)
        _statement_total("NET INCOME" if income["net_income"] >= 0 else "NET LOSS", income["net_income"])
        st.markdown("</div>", unsafe_allow_html=True)

        if st.button("📥 Export Income Statement to Excel", key="xl_income"):
            try:
                buf = generate_excel_income_statement(income, start, end, st.session_state.currency)
                st.download_button("⬇️ Download income_statement.xlsx", data=buf,
                                    file_name="income_statement.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_income")
            except Exception as exc:
                st.error(str(exc))

    with tab2:
        st.markdown(f"##### For the Period {start:%d %b %Y} to {end:%d %b %Y}")
        st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
        _statement_line("Owner's Capital, Beginning", equity["beginning_capital"])
        if abs(equity["investment"]) > 0.005:
            _statement_line("Add: Owner's Investment", equity["investment"])
        if equity["net_income"] >= 0:
            _statement_line("Add: Net Income", equity["net_income"])
        else:
            _statement_line("Less: Net Loss", equity["net_income"])
        if abs(equity["drawings"]) > 0.005:
            _statement_line("Less: Owner's Drawings", -equity["drawings"])
        _statement_total("Owner's Capital, Ending", equity["ending_capital"])
        st.markdown("</div>", unsafe_allow_html=True)

        if st.button("📥 Export Statement of Owner's Equity to Excel", key="xl_equity"):
            try:
                buf = generate_excel_owners_equity(equity, start, end, st.session_state.currency)
                st.download_button("⬇️ Download owners_equity.xlsx", data=buf,
                                    file_name="statement_of_owners_equity.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_equity")
            except Exception as exc:
                st.error(str(exc))

    with tab3:
        st.markdown(f"##### As of {end:%d %b %Y}")
        bcol1, bcol2 = st.columns(2)

        with bcol1:
            st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
            st.markdown("**Assets**")
            group_titles = {"Current Asset": "Current Assets", "Fixed Asset": "Fixed Assets", "Other Asset": "Other Assets"}
            for g in ["Current Asset", "Fixed Asset", "Other Asset"]:
                items = bs["asset_groups"].get(g, [])
                if not items:
                    continue
                st.caption(group_titles[g])
                subtotal = 0.0
                for _code, nm, amt, is_contra in items:
                    _statement_line(f"Less: {nm}" if is_contra else nm, amt, indent=True)
                    subtotal += amt
                _statement_subtotal(("Net Fixed Assets" if g == "Fixed Asset" else f"Total {group_titles[g]}"), subtotal)
                st.markdown("<br>", unsafe_allow_html=True)
            _statement_total("TOTAL ASSETS", bs["total_assets"])
            st.markdown("</div>", unsafe_allow_html=True)

        with bcol2:
            st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
            st.markdown("**Liabilities**")
            l_titles = {"Current Liability": "Current Liabilities", "Long-Term Liability": "Long-Term Liabilities"}
            any_liability = False
            for g in ["Current Liability", "Long-Term Liability"]:
                items = bs["liability_groups"].get(g, [])
                if not items:
                    continue
                any_liability = True
                st.caption(l_titles[g])
                subtotal = 0.0
                for _code, nm, amt in items:
                    _statement_line(nm, amt, indent=True)
                    subtotal += amt
                _statement_subtotal(f"Total {l_titles[g]}", subtotal)
                st.markdown("<br>", unsafe_allow_html=True)
            if not any_liability:
                st.caption("(no liabilities recorded)")
            _statement_subtotal("Total Liabilities", bs["total_liabilities"])
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("**Owner's Equity**")
            _statement_line("Owner's Capital, Ending", bs["equity"], indent=True)
            _statement_subtotal("Total Owner's Equity", bs["equity"])
            st.markdown("<br>", unsafe_allow_html=True)
            _statement_total("TOTAL LIAB. & EQUITY", bs["total_liab_equity"])
            st.markdown("</div>", unsafe_allow_html=True)

        if bs["is_balanced"]:
            st.success("✓ Assets = Liabilities + Owner's Equity — the Balance Sheet balances.")
        else:
            st.error("⚠ The Balance Sheet does not balance. This usually means an account's Type/Subtype "
                      "was changed in Chart of Accounts after transactions were already posted to it.")

        if st.button("📥 Export Balance Sheet to Excel", key="xl_bs"):
            try:
                buf = generate_excel_balance_sheet(bs, end, st.session_state.currency)
                st.download_button("⬇️ Download balance_sheet.xlsx", data=buf,
                                    file_name="balance_sheet.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_bs")
            except Exception as exc:
                st.error(str(exc))

# ============================================================================
# PAGE — Chart of Accounts
# ============================================================================

def page_chart_of_accounts():
    st.title("🗂️ Chart of Accounts")

    rows = []
    for code in sorted_account_codes():
        info = account_info(code)
        rows.append({
            "Code": code, "Name": info["name"], "Type": info["type"],
            "Subtype": info["subtype"], "Normal Balance": account_normal_balance(code),
            "Contra": "Yes" if info.get("contra") else "",
        })
    st.dataframe(pd.DataFrame(rows), hide_index=True, width="stretch")

    tab_add, tab_edit, tab_remove = st.tabs(["➕ Add Account", "✏️ Edit Account", "🗑️ Remove Account"])

    # -- Add --------------------------------------------------------------
    with tab_add:
        with st.form("add_account_form", clear_on_submit=True):
            c1, c2 = st.columns(2)
            new_code = c1.text_input("Account Code", placeholder="e.g. 106")
            new_name = c2.text_input("Account Name", placeholder="e.g. Prepaid Insurance")
            c3, c4 = st.columns(2)
            new_type = c3.selectbox("Type", ACCOUNT_TYPES, key="add_acct_type")
            new_subtype = c4.selectbox("Subtype", SUBTYPES_BY_TYPE[new_type], key=f"add_acct_subtype_{new_type}")
            new_contra = st.checkbox("This is a contra account (e.g. Accumulated Depreciation) — "
                                      "reduces its category instead of adding to it")
            submitted = st.form_submit_button("Add Account", type="primary")

        if submitted:
            code_clean = new_code.strip()
            name_clean = new_name.strip()
            if not code_clean or not name_clean:
                st.error("Both Account Code and Account Name are required.")
            elif code_clean in coa():
                st.error(f"Account code **{code_clean}** already exists.")
            else:
                st.session_state.chart_of_accounts[code_clean] = {
                    "name": name_clean, "type": new_type, "subtype": new_subtype, "contra": new_contra,
                }
                data_mgr.save_chart_of_accounts(st.session_state.chart_of_accounts)
                st.success(f"Added account {code_clean} — {name_clean}.")
                st.rerun()

    # -- Edit ---------------------------------------------------------------
    with tab_edit:
        if not coa():
            st.info("No accounts to edit yet.")
        else:
            acct_labels, acct_lookup = account_select_options()
            edit_label = st.selectbox("Account to edit", options=acct_labels, key="edit_acct_select")
            edit_code = acct_lookup[edit_label]
            current = account_info(edit_code)

            e1, e2 = st.columns(2)
            edit_name = e1.text_input("Account Name", value=current["name"], key=f"edit_name_{edit_code}")
            edit_type = e2.selectbox(
                "Type", ACCOUNT_TYPES, index=ACCOUNT_TYPES.index(current["type"]),
                key=f"edit_type_{edit_code}",
            )
            sub_options = SUBTYPES_BY_TYPE[edit_type]
            default_sub = current["subtype"] if current["subtype"] in sub_options else sub_options[0]
            edit_subtype = st.selectbox(
                "Subtype", sub_options, index=sub_options.index(default_sub),
                key=f"edit_subtype_{edit_code}_{edit_type}",
            )
            edit_contra = st.checkbox(
                "This is a contra account", value=current.get("contra", False),
                key=f"edit_contra_{edit_code}",
            )

            if st.button("💾 Save Account Changes", type="primary"):
                st.session_state.chart_of_accounts[edit_code] = {
                    "name": edit_name.strip() or current["name"],
                    "type": edit_type, "subtype": edit_subtype, "contra": edit_contra,
                }
                data_mgr.save_chart_of_accounts(st.session_state.chart_of_accounts)
                st.success(f"Account {edit_code} updated.")
                st.rerun()
            st.caption("Changing an account's Type/Subtype after transactions have already been "
                       "posted to it will change how it's classified in future reports.")

    # -- Remove ---------------------------------------------------------------
    with tab_remove:
        if not coa():
            st.info("No accounts to remove.")
        else:
            acct_labels, acct_lookup = account_select_options()
            remove_label = st.selectbox("Account to remove", options=acct_labels, key="remove_acct_select")
            remove_code = acct_lookup[remove_label]
            in_use = (journal_df()["Post_Ref"] == remove_code).any() if not journal_df().empty else False

            if in_use:
                st.warning(f"Account **{remove_code} — {account_name(remove_code)}** has transactions posted "
                           "to it and can't be removed. Delete those transactions first if you really need to.")
            else:
                if st.session_state.confirm_remove_acct == remove_code:
                    st.warning(f"Remove account **{remove_code} — {account_name(remove_code)}** permanently?")
                    rc1, rc2 = st.columns(2)
                    if rc1.button("✅ Yes, remove it", type="primary"):
                        del st.session_state.chart_of_accounts[remove_code]
                        data_mgr.save_chart_of_accounts(st.session_state.chart_of_accounts)
                        st.session_state.confirm_remove_acct = None
                        st.success("Account removed.")
                        st.rerun()
                    if rc2.button("Cancel", key="cancel_remove_acct"):
                        st.session_state.confirm_remove_acct = None
                        st.rerun()
                else:
                    if st.button("🗑️ Remove This Account"):
                        st.session_state.confirm_remove_acct = remove_code
                        st.rerun()

# ============================================================================
# PAGE — Export Full Cycle
# ============================================================================

def page_export_full_cycle():
    st.title("📦 Export Full Cycle")
    st.caption(
        "One Excel workbook covering the complete accounting cycle — from the raw "
        "General Journal all the way through the Financial Statements — ready to open in Excel."
    )

    df = journal_df()
    if df.empty:
        st.info("No transactions recorded yet. Record entries in **Input Journal** first.")
        return

    inception = df["Date"].min()
    latest = df["Date"].max()

    st.markdown('<div class="ledger-card">', unsafe_allow_html=True)
    st.markdown("**This workbook includes, one sheet per report, in order:**")
    st.markdown(
        "1. **General Journal** — every posted transaction, chronologically  \n"
        "2. **General Ledger** — every account with activity, with running balances  \n"
        "3. **Trial Balance** — as of the latest transaction date  \n"
        "4. **Income Statement** — full period, inception to latest date  \n"
        "5. **Statement of Owner's Equity** — full period  \n"
        "6. **Balance Sheet** — as of the latest transaction date"
    )
    st.markdown("</div>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    c1.metric("Period Covered", f"{inception:%d %b %Y} – {latest:%d %b %Y}")
    c2.metric("Transactions", df["Transaction_ID"].nunique())
    c3.metric("Accounts With Activity", df["Post_Ref"].nunique())

    if st.button("📥 Generate Full Cycle Workbook", type="primary", width="stretch"):
        try:
            buf = generate_excel_full_cycle(df, st.session_state.currency)
            st.success("Workbook ready — download below.")
            st.download_button(
                "⬇️ Download full_accounting_cycle.xlsx",
                data=buf,
                file_name=f"full_accounting_cycle_{latest:%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        except Exception as exc:
            st.error(str(exc))

# ============================================================================
# MAIN
# ============================================================================

PAGE_FUNCS = {
    "Input Journal": page_input_journal,
    "View Journal": page_view_journal,
    "View Ledger": page_view_ledger,
    "Trial Balance": page_trial_balance,
    "Financial Statements": page_financial_statements,
    "Export Full Cycle": page_export_full_cycle,
    "Chart of Accounts": page_chart_of_accounts,
}

def main():
    render_sidebar()
    PAGE_FUNCS.get(st.session_state.nav_page, page_input_journal)()

if __name__ == "__main__":
    main()